#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Trade Journal Generator for Mercurio AI
--------------------------------------
Ce script génère un carnet de bord complet à partir des logs et rapports de trading,
puis l'exporte au format Excel (.xlsx) avec des graphiques et analyses.

Utilisation:
    python scripts/trade_journal.py --output carnet_de_bord.xlsx
"""

import os
import sys
import json
import glob
import logging
import argparse
import re
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart, PieChart
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

# Ajouter le répertoire parent au path pour pouvoir importer les modules du projet
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.append(project_root)

# Configuration du logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger('trade_journal')

# Couleurs pour le style Excel
HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
MONEY_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
PCT_FORMAT = '0.00%'
DATE_FORMAT = 'yyyy-mm-dd hh:mm:ss'
BORDER = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

class TradeJournalGenerator:
    """Générateur de carnet de bord pour les transactions de trading."""
    
    def __init__(self, output_file=None):
        """
        Initialise le générateur de carnet de bord.
        
        Args:
            output_file: Chemin du fichier Excel à générer
        """
        self.output_file = output_file or f"trade_journal_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.logs_dir = os.path.join(project_root, "logs")
        self.outputs_dir = os.path.join(project_root, "outputs")
        
        # Dataframes pour stocker les données
        self.transactions_df = pd.DataFrame()
        self.daily_performance_df = pd.DataFrame()
        self.strategy_performance_df = pd.DataFrame()
        self.symbol_performance_df = pd.DataFrame()
        self.trade_stats_df = pd.DataFrame()
    
    def collect_data(self):
        """Collecte toutes les données de trading depuis les différentes sources."""
        logger.info("Collecte des données de trading...")
        
        # 1. Analyser les fichiers de rapports dans le dossier outputs
        self._parse_report_files()
        
        # 2. Analyser les fichiers de logs
        self._parse_log_files()
        
        # 3. Générer des statistiques agrégées
        self._generate_statistics()
        
        logger.info("Collecte des données terminée.")
    
    def _parse_report_files(self):
        """Analyse les fichiers de rapports JSON dans le dossier outputs."""
        report_files = glob.glob(os.path.join(self.outputs_dir, "*.json"))
        logger.info(f"Analyse de {len(report_files)} fichiers de rapports...")
        
        all_trades = []
        strategy_performances = []
        
        for file_path in report_files:
            try:
                with open(file_path, 'r') as f:
                    report = json.load(f)
                
                # Extraire les informations générales du rapport
                strategy = report.get("strategy")
                start_time = datetime.fromisoformat(report.get("start_time", "").replace("Z", "+00:00")) if "start_time" in report else None
                end_time = datetime.fromisoformat(report.get("end_time", "").replace("Z", "+00:00")) if "end_time" in report else None
                
                # Enregistrer les performances de la stratégie
                strategy_data = {
                    "strategy": strategy,
                    "start_time": start_time,
                    "end_time": end_time,
                    "symbols_count": report.get("symbols_count", 0),
                    "positions_opened": report.get("positions_opened", 0),
                    "initial_capital": report.get("settings", {}).get("capital", 0),
                    "allocation_per_trade": report.get("settings", {}).get("allocation_per_trade", 0),
                    "profit_target": report.get("settings", {}).get("profit_target", 0),
                    "stop_loss": report.get("settings", {}).get("stop_loss", 0),
                    "max_symbols": report.get("settings", {}).get("max_symbols", 0),
                    "use_custom_symbols": report.get("settings", {}).get("use_custom_symbols", False),
                }
                
                # Ajouter les métriques de performance si disponibles
                if "total_return" in report:
                    strategy_data["total_return"] = report.get("total_return", 0)
                    strategy_data["total_return_pct"] = report.get("total_return_pct", 0) / 100 if "total_return_pct" in report else 0
                    strategy_data["win_rate"] = report.get("win_rate", 0)
                    strategy_data["profitable_trades"] = report.get("profitable_trades", 0)
                    strategy_data["losing_trades"] = report.get("losing_trades", 0)
                    strategy_data["total_trades"] = report.get("total_trades", 0)
                
                strategy_performances.append(strategy_data)
                
                # Extraire les transactions individuelles
                for trade in report.get("trades", []):
                    trade_data = {
                        "strategy": strategy,
                        "symbol": trade.get("symbol"),
                        "entry_time": datetime.fromisoformat(trade.get("entry_time", "").replace("Z", "+00:00")) if "entry_time" in trade else None,
                        "exit_time": datetime.fromisoformat(trade.get("exit_time", "").replace("Z", "+00:00")) if "exit_time" in trade else None,
                        "entry_price": trade.get("entry_price", 0),
                        "exit_price": trade.get("exit_price", 0),
                        "quantity": trade.get("quantity", 0),
                        "side": trade.get("side", ""),
                        "profit_loss": trade.get("profit_loss", 0),
                        "profit_loss_pct": trade.get("profit_loss_pct", 0) / 100 if "profit_loss_pct" in trade else 0,
                    }
                    all_trades.append(trade_data)
            except Exception as e:
                logger.error(f"Erreur lors de l'analyse du fichier {file_path}: {e}")
        
        # Convertir en DataFrames si des données ont été trouvées
        if all_trades:
            self.transactions_df = pd.DataFrame(all_trades)
            # Convertir les dates en datetime si nécessaire
            for date_col in ['entry_time', 'exit_time']:
                if date_col in self.transactions_df.columns:
                    self.transactions_df[date_col] = pd.to_datetime(self.transactions_df[date_col])
        
        if strategy_performances:
            self.strategy_performance_df = pd.DataFrame(strategy_performances)
            # Convertir les dates en datetime si nécessaire
            for date_col in ['start_time', 'end_time']:
                if date_col in self.strategy_performance_df.columns:
                    self.strategy_performance_df[date_col] = pd.to_datetime(self.strategy_performance_df[date_col])
    
    def _parse_log_files(self):
        """Analyse les fichiers de logs pour extraire des informations sur les transactions."""
        log_files = glob.glob(os.path.join(self.logs_dir, "*.log"))
        logger.info(f"Analyse de {len(log_files)} fichiers de logs...")
        
        # Patterns regex pour extraire les informations de trading des logs
        buy_pattern = re.compile(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}).*Ordre d\'achat placé pour ([A-Z]+):\s*(\d+)\s*actions à ~\$([\d.]+)')
        sell_pattern = re.compile(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}).*Ordre de vente placé pour ([A-Z]+):\s*([\d.]+)\s*actions à ~\$([\d.]+)')
        position_pattern = re.compile(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}).*Position pour ([A-Z]+): ([\d.]+) actions à prix moyen \$([\d.]+)')
        perf_pattern = re.compile(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}).*Performance du jour: \$([\d.-]+) \(([\d.-]+)%\)')
        
        # Listes pour stocker les données extraites
        log_trades = []
        daily_performances = []
        
        for log_file in log_files:
            try:
                with open(log_file, 'r', encoding='utf-8', errors='ignore') as f:
                    log_content = f.readlines()
                
                # Extraire la stratégie du nom du fichier log si possible
                strategy_match = re.search(r'([a-z_]+)_trader\.log', os.path.basename(log_file))
                strategy = strategy_match.group(1) if strategy_match else "unknown"
                
                for line in log_content:
                    # Analyser les ordres d'achat
                    buy_match = buy_pattern.search(line)
                    if buy_match:
                        timestamp, symbol, qty, price = buy_match.groups()
                        log_trades.append({
                            "timestamp": timestamp,
                            "symbol": symbol,
                            "quantity": float(qty),
                            "price": float(price),
                            "action": "BUY",
                            "strategy": strategy
                        })
                        continue
                    
                    # Analyser les ordres de vente
                    sell_match = sell_pattern.search(line)
                    if sell_match:
                        timestamp, symbol, qty, price = sell_match.groups()
                        log_trades.append({
                            "timestamp": timestamp,
                            "symbol": symbol,
                            "quantity": float(qty),
                            "price": float(price),
                            "action": "SELL",
                            "strategy": strategy
                        })
                        continue
                    
                    # Analyser les performances journalières
                    perf_match = perf_pattern.search(line)
                    if perf_match:
                        timestamp, profit_usd, profit_pct = perf_match.groups()
                        daily_performances.append({
                            "date": datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S").date(),
                            "profit_usd": float(profit_usd),
                            "profit_pct": float(profit_pct) / 100,  # Convertir en décimal
                            "strategy": strategy
                        })
            except Exception as e:
                logger.error(f"Erreur lors de l'analyse du fichier {log_file}: {e}")
        
        # Convertir en DataFrames si des données ont été trouvées
        if log_trades:
            log_trades_df = pd.DataFrame(log_trades)
            log_trades_df['timestamp'] = pd.to_datetime(log_trades_df['timestamp'])
            
            # Fusionner avec les transactions existantes si nécessaire
            if not self.transactions_df.empty:
                # Utilisez des clés appropriées pour éviter les doublons
                self.transactions_df = pd.concat([self.transactions_df, log_trades_df])
            else:
                self.transactions_df = log_trades_df
        
        if daily_performances:
            daily_perf_df = pd.DataFrame(daily_performances)
            
            # Fusionner avec les performances existantes si nécessaire
            if not self.daily_performance_df.empty:
                self.daily_performance_df = pd.concat([self.daily_performance_df, daily_perf_df])
            else:
                self.daily_performance_df = daily_perf_df
    
    def _generate_statistics(self):
        """Génère des statistiques agrégées à partir des données collectées."""
        logger.info("Génération des statistiques...")
        
        # Skip si aucune donnée n'est disponible
        if self.transactions_df.empty:
            logger.warning("Aucune transaction trouvée pour générer des statistiques")
            return
        
        # 1. Statistiques par symbole
        try:
            if 'symbol' in self.transactions_df.columns:
                symbol_stats = self.transactions_df.groupby('symbol').agg(
                    total_trades=pd.NamedAgg(column='symbol', aggfunc='count'),
                    avg_price=pd.NamedAgg(column='price', aggfunc='mean')
                ).reset_index()
                
                # Calculer les profits par symbole si possible
                if all(col in self.transactions_df.columns for col in ['profit_loss', 'symbol']):
                    profit_by_symbol = self.transactions_df.groupby('symbol')['profit_loss'].sum().reset_index()
                    symbol_stats = symbol_stats.merge(profit_by_symbol, on='symbol', how='left')
                
                self.symbol_performance_df = symbol_stats
        except Exception as e:
            logger.error(f"Erreur lors de la génération des statistiques par symbole: {e}")
        
        # 2. Statistiques globales de trading
        try:
            stats = {
                'date_generated': datetime.now(),
                'total_transactions': len(self.transactions_df),
            }
            
            # Calculs supplémentaires si les colonnes nécessaires sont disponibles
            if 'profit_loss' in self.transactions_df.columns:
                stats['total_profit_loss'] = self.transactions_df['profit_loss'].sum()
                stats['avg_profit_loss'] = self.transactions_df['profit_loss'].mean()
                stats['max_profit'] = self.transactions_df['profit_loss'].max()
                stats['max_loss'] = self.transactions_df['profit_loss'].min()
                
                # Calculer le taux de réussite
                profitable_trades = (self.transactions_df['profit_loss'] > 0).sum()
                losing_trades = (self.transactions_df['profit_loss'] < 0).sum()
                stats['profitable_trades'] = profitable_trades
                stats['losing_trades'] = losing_trades
                stats['win_rate'] = profitable_trades / len(self.transactions_df) if len(self.transactions_df) > 0 else 0
            
            # Nombre de jours de trading
            if 'entry_time' in self.transactions_df.columns:
                unique_days = self.transactions_df['entry_time'].dt.date.nunique()
                stats['trading_days'] = unique_days
                stats['avg_trades_per_day'] = len(self.transactions_df) / unique_days if unique_days > 0 else 0
            
            self.trade_stats_df = pd.DataFrame([stats])
        except Exception as e:
            logger.error(f"Erreur lors de la génération des statistiques globales: {e}")
    
    def generate_excel_report(self):
        """
        Génère le rapport Excel complet avec toutes les feuilles et graphiques.
        """
        logger.info(f"Génération du rapport Excel: {self.output_file}")
        
        # Créer un nouveau workbook
        wb = Workbook()
        # Supprimer la feuille par défaut
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Créer les différentes feuilles du rapport
        self._create_summary_sheet(wb)
        self._create_transactions_sheet(wb)
        self._create_daily_performance_sheet(wb)
        self._create_symbol_performance_sheet(wb)
        self._create_strategy_performance_sheet(wb)
        
        # Sauvegarder le workbook
        try:
            wb.save(self.output_file)
            logger.info(f"Rapport Excel généré avec succès: {self.output_file}")
            return True
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde du rapport Excel: {e}")
            return False
    
    def _create_summary_sheet(self, wb):
        """
        Crée la feuille de résumé avec les statistiques globales.
        """
        ws = wb.create_sheet("Résumé")
        
        # Titre principal
        ws['A1'] = "CARNET DE BORD DE TRADING - MERCURIO AI"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Date de génération
        ws['A3'] = "Date de génération:"
        ws['B3'] = datetime.now()
        ws['B3'].number_format = DATE_FORMAT
        
        # Ajouter les statistiques globales si disponibles
        if not self.trade_stats_df.empty:
            row = 5
            ws[f'A{row}'] = "STATISTIQUES GLOBALES"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 2
            
            # Extraction des statistiques
            stats = self.trade_stats_df.iloc[0].to_dict()
            
            # Ajout des statistiques principales
            metrics = [
                ("Nombre total de transactions", stats.get("total_transactions", 0), None),
                ("Jours de trading", stats.get("trading_days", 0), None),
                ("Transactions par jour (moyenne)", stats.get("avg_trades_per_day", 0), None),
                ("Taux de réussite", stats.get("win_rate", 0), PCT_FORMAT),
                ("Transactions profitables", stats.get("profitable_trades", 0), None),
                ("Transactions perdantes", stats.get("losing_trades", 0), None),
                ("Profit/Perte total", stats.get("total_profit_loss", 0), MONEY_FORMAT),
                ("Profit/Perte moyen par trade", stats.get("avg_profit_loss", 0), MONEY_FORMAT),
                ("Profit maximum", stats.get("max_profit", 0), MONEY_FORMAT),
                ("Perte maximum", stats.get("max_loss", 0), MONEY_FORMAT),
            ]
            
            for i, (label, value, fmt) in enumerate(metrics):
                ws[f'A{row + i}'] = label
                ws[f'B{row + i}'] = value
                if fmt:
                    ws[f'B{row + i}'].number_format = fmt
            
            # Ajouter un petit graphique résumant les profits/pertes si les données sont disponibles
            if not self.daily_performance_df.empty and 'profit_usd' in self.daily_performance_df.columns:
                self._add_summary_profit_chart(wb, ws, row + len(metrics) + 2)
        else:
            ws['A5'] = "Aucune statistique disponible. Exécutez d'abord le trading pour générer des données."
            ws['A5'].font = Font(italic=True)
    
    def _add_summary_profit_chart(self, wb, ws, start_row):
        """
        Ajoute un graphique de profits cumulés à la feuille de résumé.
        """
        try:
            # Préparer les données pour le graphique
            if 'date' in self.daily_performance_df.columns and 'profit_usd' in self.daily_performance_df.columns:
                # Trier par date
                perf_df = self.daily_performance_df.sort_values('date')
                # Calculer le profit cumulé
                perf_df['cumulative_profit'] = perf_df['profit_usd'].cumsum()
                
                # Ajouter un titre
                ws[f'A{start_row}'] = "ÉVOLUTION DU PROFIT CUMULÉ"
                ws[f'A{start_row}'].font = Font(bold=True, size=14)
                ws.merge_cells(f'A{start_row}:H{start_row}')
                ws[f'A{start_row}'].alignment = Alignment(horizontal='center')
                
                # Ajouter les données pour le graphique
                ws[f'A{start_row + 2}'] = "Date"
                ws[f'B{start_row + 2}'] = "Profit cumulé"
                
                for i, (_, row) in enumerate(perf_df.iterrows()):
                    ws[f'A{start_row + 3 + i}'] = row['date']
                    ws[f'B{start_row + 3 + i}'] = row['cumulative_profit']
                    ws[f'B{start_row + 3 + i}'].number_format = MONEY_FORMAT
                
                # Créer le graphique
                chart = LineChart()
                chart.title = "Evolution du profit cumulé"
                chart.style = 2
                chart.x_axis.title = "Date"
                chart.y_axis.title = "Profit cumulé ($)"
                
                # Ajouter les données au graphique
                data = Reference(ws, min_col=2, min_row=start_row + 2, max_row=start_row + 2 + len(perf_df))
                cats = Reference(ws, min_col=1, min_row=start_row + 3, max_row=start_row + 2 + len(perf_df))
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                # Placer le graphique
                ws.add_chart(chart, f'D{start_row + 2}')
        except Exception as e:
            logger.error(f"Erreur lors de la création du graphique de profits: {e}")
    
    def _create_transactions_sheet(self, wb):
        """
        Crée la feuille contenant toutes les transactions.
        """
        ws = wb.create_sheet("Transactions")
        
        # Titre de la feuille
        ws['A1'] = "JOURNAL DES TRANSACTIONS"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        if not self.transactions_df.empty:
            # Adapter les noms de colonnes si nécessaire
            df_to_display = self.transactions_df.copy()
            
            # Unifier les noms de colonnes (différentes sources peuvent avoir des noms différents)
            col_mapping = {
                'entry_time': 'Date d\'entrée',
                'exit_time': 'Date de sortie',
                'timestamp': 'Date',
                'symbol': 'Symbole',
                'quantity': 'Quantité',
                'price': 'Prix',
                'entry_price': 'Prix d\'entrée',
                'exit_price': 'Prix de sortie',
                'side': 'Direction',
                'action': 'Action',
                'profit_loss': 'Profit/Perte',
                'profit_loss_pct': '% Profit/Perte',
                'strategy': 'Stratégie'
            }
            
            # Renommer les colonnes si elles existent
            for old_col, new_col in col_mapping.items():
                if old_col in df_to_display.columns:
                    df_to_display.rename(columns={old_col: new_col}, inplace=True)
            
            # Trier par date si possible
            date_cols = [col for col in df_to_display.columns if 'date' in col.lower()]
            if date_cols:
                df_to_display.sort_values(date_cols[0], ascending=False, inplace=True)
            
            # Ajouter les données à la feuille
            for r_idx, row in enumerate(dataframe_to_rows(df_to_display, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=r_idx + 3, column=c_idx + 1, value=value)
                    
                    # Appliquer un style aux cellules d'en-tête
                    if r_idx == 0:
                        cell.font = HEADER_FONT
                        cell.fill = HEADER_FILL
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = BORDER
                    else:
                        # Appliquer des formats spécifiques selon le type de données
                        if isinstance(value, (int, float)) and 'prix' in ws.cell(row=3, column=c_idx + 1).value.lower():
                            cell.number_format = MONEY_FORMAT
                        elif isinstance(value, (int, float)) and 'profit' in ws.cell(row=3, column=c_idx + 1).value.lower():
                            cell.number_format = MONEY_FORMAT
                        elif isinstance(value, (int, float)) and '%' in ws.cell(row=3, column=c_idx + 1).value:
                            cell.number_format = PCT_FORMAT
                        elif isinstance(value, datetime):
                            cell.number_format = DATE_FORMAT
                            
                        # Colorer les profits/pertes
                        if 'profit' in ws.cell(row=3, column=c_idx + 1).value.lower():
                            if isinstance(value, (int, float)):
                                if value > 0:
                                    cell.font = Font(color="00A952")  # Vert
                                elif value < 0:
                                    cell.font = Font(color="D3242C")  # Rouge
            
            # Ajuster la largeur des colonnes
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2
        else:
            ws['A3'] = "Aucune transaction disponible."
            ws['A3'].font = Font(italic=True)
    
    def _create_daily_performance_sheet(self, wb):
        """
        Crée la feuille de performance journalière.
        """
        ws = wb.create_sheet("Performance Journalière")
        
        # Titre de la feuille
        ws['A1'] = "PERFORMANCE JOURNALIÈRE"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        if not self.daily_performance_df.empty:
            # Préparer les données
            df_to_display = self.daily_performance_df.copy()
            
            # Renommer les colonnes si nécessaire
            col_mapping = {
                'date': 'Date',
                'total_trades': 'Nombre de transactions',
                'profitable_trades': 'Transactions profitables',
                'losing_trades': 'Transactions perdantes',
                'win_rate': 'Taux de réussite',
                'profit_usd': 'Profit/Perte ($)',
                'profit_pct': 'Profit/Perte (%)',
                'cumulative_profit': 'Profit cumulé ($)'
            }
            
            for old_col, new_col in col_mapping.items():
                if old_col in df_to_display.columns:
                    df_to_display.rename(columns={old_col: new_col}, inplace=True)
            
            # Trier par date
            if 'Date' in df_to_display.columns:
                df_to_display.sort_values('Date', ascending=False, inplace=True)
            
            # Ajouter les données à la feuille
            for r_idx, row in enumerate(dataframe_to_rows(df_to_display, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=r_idx + 3, column=c_idx + 1, value=value)
                    
                    # Appliquer des styles à l'en-tête
                    if r_idx == 0:
                        cell.font = HEADER_FONT
                        cell.fill = HEADER_FILL
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = BORDER
                    else:
                        # Formater les cellules selon leur contenu
                        header_value = ws.cell(row=3, column=c_idx + 1).value
                        if isinstance(value, (int, float)) and 'profit' in str(header_value).lower() and '$' in str(header_value):
                            cell.number_format = MONEY_FORMAT
                            if value > 0:
                                cell.font = Font(color="00A952")  # Vert
                            elif value < 0:
                                cell.font = Font(color="D3242C")  # Rouge
                        elif isinstance(value, (int, float)) and '%' in str(header_value):
                            cell.number_format = PCT_FORMAT
                            if 'profit' in str(header_value).lower():
                                if value > 0:
                                    cell.font = Font(color="00A952")
                                elif value < 0:
                                    cell.font = Font(color="D3242C")
                        elif isinstance(value, datetime):
                            cell.number_format = DATE_FORMAT
            
            # Ajuster la largeur des colonnes
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2
            
            # Ajouter un graphique de performance journalière si les données sont disponibles
            self._add_daily_performance_chart(wb, ws, len(df_to_display) + 5)
        else:
            ws['A3'] = "Aucune donnée de performance journalière disponible."
            ws['A3'].font = Font(italic=True)
    
    def _add_daily_performance_chart(self, wb, ws, start_row):
        """
        Ajoute un graphique de performance journalière.
        """
        try:
            if 'Date' in [cell.value for cell in ws[3]] and any('Profit' in str(cell.value) for cell in ws[3]):
                # Trouver l'index de la colonne de date et de profit
                date_col = None
                profit_col = None
                for col in range(1, ws.max_column + 1):
                    header = ws.cell(row=3, column=col).value
                    if header == 'Date':
                        date_col = col
                    elif 'Profit/Perte ($)' in str(header):
                        profit_col = col
                
                if date_col is not None and profit_col is not None:
                    # Titre pour la section du graphique
                    ws[f'A{start_row}'] = "GRAPHIQUE DE PERFORMANCE JOURNALIÈRE"
                    ws[f'A{start_row}'].font = Font(bold=True, size=14)
                    ws.merge_cells(f'A{start_row}:F{start_row}')
                    ws[f'A{start_row}'].alignment = Alignment(horizontal='center')
                    
                    # Créer le graphique
                    chart = BarChart()
                    chart.type = "col"
                    chart.style = 10
                    chart.title = "Performance journalière"
                    chart.x_axis.title = "Date"
                    chart.y_axis.title = "Profit/Perte ($)"
                    
                    # Références pour les données
                    data = Reference(ws, min_col=profit_col, min_row=3, max_row=ws.max_row)
                    cats = Reference(ws, min_col=date_col, min_row=4, max_row=ws.max_row)
                    
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    
                    # Ajouter le graphique à la feuille
                    ws.add_chart(chart, f'A{start_row + 2}')
        except Exception as e:
            logger.error(f"Erreur lors de la création du graphique de performance journalière: {e}")
    
    def _create_symbol_performance_sheet(self, wb):
        """
        Crée la feuille de performance par symbole.
        """
        ws = wb.create_sheet("Performance par Symbole")
        
        # Titre de la feuille
        ws['A1'] = "PERFORMANCE PAR SYMBOLE"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        if not self.symbol_performance_df.empty:
            # Préparer les données
            df_to_display = self.symbol_performance_df.copy()
            
            # Renommer les colonnes si nécessaire
            col_mapping = {
                'symbol': 'Symbole',
                'total_trades': 'Nombre de transactions',
                'avg_price': 'Prix moyen',
                'profit_loss': 'Profit/Perte',
                'win_rate': 'Taux de réussite'
            }
            
            for old_col, new_col in col_mapping.items():
                if old_col in df_to_display.columns:
                    df_to_display.rename(columns={old_col: new_col}, inplace=True)
            
            # Trier par profit si disponible, sinon par nombre de transactions
            if 'Profit/Perte' in df_to_display.columns:
                df_to_display.sort_values('Profit/Perte', ascending=False, inplace=True)
            elif 'Nombre de transactions' in df_to_display.columns:
                df_to_display.sort_values('Nombre de transactions', ascending=False, inplace=True)
            
            # Ajouter les données à la feuille
            for r_idx, row in enumerate(dataframe_to_rows(df_to_display, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=r_idx + 3, column=c_idx + 1, value=value)
                    
                    # Appliquer des styles à l'en-tête
                    if r_idx == 0:
                        cell.font = HEADER_FONT
                        cell.fill = HEADER_FILL
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = BORDER
                    else:
                        # Formater les cellules selon leur contenu
                        header_value = ws.cell(row=3, column=c_idx + 1).value
                        if isinstance(value, (int, float)) and 'prix' in str(header_value).lower():
                            cell.number_format = MONEY_FORMAT
                        elif isinstance(value, (int, float)) and 'profit' in str(header_value).lower():
                            cell.number_format = MONEY_FORMAT
                            if value > 0:
                                cell.font = Font(color="00A952")  # Vert
                            elif value < 0:
                                cell.font = Font(color="D3242C")  # Rouge
                        elif isinstance(value, (int, float)) and '%' in str(header_value):
                            cell.number_format = PCT_FORMAT
            
            # Ajuster la largeur des colonnes
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2
            
            # Ajouter un graphique de performance par symbole
            self._add_symbol_performance_chart(wb, ws, len(df_to_display) + 5)
        else:
            ws['A3'] = "Aucune donnée de performance par symbole disponible."
            ws['A3'].font = Font(italic=True)
    
    def _add_symbol_performance_chart(self, wb, ws, start_row):
        """
        Ajoute un graphique de performance par symbole.
        """
        try:
            if 'Symbole' in [cell.value for cell in ws[3]] and any('Profit' in str(cell.value) for cell in ws[3]):
                # Trouver l'index des colonnes de symbole et de profit
                symbol_col = None
                profit_col = None
                for col in range(1, ws.max_column + 1):
                    header = ws.cell(row=3, column=col).value
                    if header == 'Symbole':
                        symbol_col = col
                    elif header == 'Profit/Perte':
                        profit_col = col
                
                if symbol_col is not None and profit_col is not None:
                    # Titre pour la section du graphique
                    ws[f'A{start_row}'] = "TOP SYMBOLES PAR PERFORMANCE"
                    ws[f'A{start_row}'].font = Font(bold=True, size=14)
                    ws.merge_cells(f'A{start_row}:F{start_row}')
                    ws[f'A{start_row}'].alignment = Alignment(horizontal='center')
                    
                    # Prendre les 10 premiers symboles (ou moins s'il y en a moins de 10)
                    max_symbols = min(10, ws.max_row - 3)
                    
                    # Créer le graphique
                    chart = BarChart()
                    chart.type = "col"
                    chart.style = 10
                    chart.title = "Performance par symbole"
                    chart.x_axis.title = "Symbole"
                    chart.y_axis.title = "Profit/Perte ($)"
                    
                    # Références pour les données
                    data = Reference(ws, min_col=profit_col, min_row=3, max_row=3 + max_symbols)
                    cats = Reference(ws, min_col=symbol_col, min_row=4, max_row=3 + max_symbols)
                    
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    
                    # Personnaliser la couleur des barres
                    # Fonction qui détermine la couleur basée sur la valeur
                    for i in range(max_symbols):
                        profit = ws.cell(row=4 + i, column=profit_col).value
                        if profit > 0:
                            chart.series[0][i].graphicalProperties.solidFill = "00A952"  # Vert
                        else:
                            chart.series[0][i].graphicalProperties.solidFill = "D3242C"  # Rouge
                    
                    # Ajouter le graphique à la feuille
                    ws.add_chart(chart, f'A{start_row + 2}')
        except Exception as e:
            logger.error(f"Erreur lors de la création du graphique de performance par symbole: {e}")
    
    def _create_strategy_performance_sheet(self, wb):
        """
        Crée la feuille de performance par stratégie.
        """
        ws = wb.create_sheet("Performance par Stratégie")
        
        # Titre de la feuille
        ws['A1'] = "PERFORMANCE PAR STRATÉGIE"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Vérifier si nous avons des données de stratégie
        if 'strategy' in self.transactions_df.columns or 'stratégie' in self.transactions_df.columns:
            # Créer un dataframe de performance par stratégie
            strategy_col = 'strategy' if 'strategy' in self.transactions_df.columns else 'stratégie'
            
            try:
                # Regrouper par stratégie et calculer les métriques
                strategy_performance = self.transactions_df.groupby(strategy_col).agg(
                    total_trades=pd.NamedAgg(column=strategy_col, aggfunc='count')
                ).reset_index()
                
                # Ajouter la colonne de profit/perte si disponible
                if 'profit_loss' in self.transactions_df.columns:
                    profit_by_strategy = self.transactions_df.groupby(strategy_col)['profit_loss'].sum().reset_index()
                    strategy_performance = strategy_performance.merge(profit_by_strategy, on=strategy_col, how='left')
                    
                    # Calculer le taux de réussite par stratégie
                    win_rates = []
                    for strategy in strategy_performance[strategy_col]:
                        strategy_trades = self.transactions_df[self.transactions_df[strategy_col] == strategy]
                        profitable_trades = (strategy_trades['profit_loss'] > 0).sum()
                        win_rate = profitable_trades / len(strategy_trades) if len(strategy_trades) > 0 else 0
                        win_rates.append(win_rate)
                    
                    strategy_performance['win_rate'] = win_rates
                
                # Préparer les données pour l'affichage
                df_to_display = strategy_performance.copy()
                
                # Renommer les colonnes
                col_mapping = {
                    strategy_col: 'Stratégie',
                    'total_trades': 'Nombre de transactions',
                    'profit_loss': 'Profit/Perte',
                    'win_rate': 'Taux de réussite'
                }
                
                for old_col, new_col in col_mapping.items():
                    if old_col in df_to_display.columns:
                        df_to_display.rename(columns={old_col: new_col}, inplace=True)
                
                # Trier par profit si disponible
                if 'Profit/Perte' in df_to_display.columns:
                    df_to_display.sort_values('Profit/Perte', ascending=False, inplace=True)
                
                # Ajouter les données à la feuille
                for r_idx, row in enumerate(dataframe_to_rows(df_to_display, index=False, header=True)):
                    for c_idx, value in enumerate(row):
                        cell = ws.cell(row=r_idx + 3, column=c_idx + 1, value=value)
                        
                        # Appliquer des styles à l'en-tête
                        if r_idx == 0:
                            cell.font = HEADER_FONT
                            cell.fill = HEADER_FILL
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = BORDER
                        else:
                            # Formater les cellules selon leur contenu
                            header_value = ws.cell(row=3, column=c_idx + 1).value
                            if isinstance(value, (int, float)) and 'profit' in str(header_value).lower():
                                cell.number_format = MONEY_FORMAT
                                if value > 0:
                                    cell.font = Font(color="00A952")  # Vert
                                elif value < 0:
                                    cell.font = Font(color="D3242C")  # Rouge
                            elif isinstance(value, (int, float)) and 'taux' in str(header_value).lower():
                                cell.number_format = PCT_FORMAT
                
                # Ajuster la largeur des colonnes
                for column_cells in ws.columns:
                    max_length = 0
                    column = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column].width = max_length + 2
                
                # Ajouter un graphique de performance par stratégie
                if len(df_to_display) > 0:
                    self._add_strategy_performance_chart(wb, ws, len(df_to_display) + 5)
            except Exception as e:
                logger.error(f"Erreur lors de la création de la feuille de performance par stratégie: {e}")
                ws['A3'] = f"Erreur lors de l'analyse des données de stratégie: {e}"
                ws['A3'].font = Font(italic=True)
        else:
            ws['A3'] = "Aucune donnée de stratégie disponible dans les transactions."
            ws['A3'].font = Font(italic=True)
    
    def _add_strategy_performance_chart(self, wb, ws, start_row):
        """
        Ajoute un graphique de performance par stratégie.
        """
        try:
            if 'Stratégie' in [cell.value for cell in ws[3]] and any('Profit' in str(cell.value) for cell in ws[3]):
                # Trouver l'index des colonnes de stratégie et de profit
                strategy_col = None
                profit_col = None
                for col in range(1, ws.max_column + 1):
                    header = ws.cell(row=3, column=col).value
                    if header == 'Stratégie':
                        strategy_col = col
                    elif header == 'Profit/Perte':
                        profit_col = col
                
                if strategy_col is not None and profit_col is not None:
                    # Titre pour la section du graphique
                    ws[f'A{start_row}'] = "PERFORMANCE PAR STRATÉGIE"
                    ws[f'A{start_row}'].font = Font(bold=True, size=14)
                    ws.merge_cells(f'A{start_row}:F{start_row}')
                    ws[f'A{start_row}'].alignment = Alignment(horizontal='center')
                    
                    # Créer le graphique
                    chart = PieChart()
                    chart.title = "Répartition des profits par stratégie"
                    
                    # Références pour les données
                    data = Reference(ws, min_col=profit_col, min_row=4, max_row=ws.max_row)
                    cats = Reference(ws, min_col=strategy_col, min_row=4, max_row=ws.max_row)
                    
                    chart.add_data(data)
                    chart.set_categories(cats)
                    
                    # Ajouter le graphique à la feuille
                    ws.add_chart(chart, f'A{start_row + 2}')
        except Exception as e:
            logger.error(f"Erreur lors de la création du graphique de performance par stratégie: {e}")


def parse_arguments():
    """
    Parse les arguments de ligne de commande.
    
    Returns:
        Les arguments parsés.
    """
    parser = argparse.ArgumentParser(description='Générateur de carnet de bord de trading Mercurio AI')
    parser.add_argument('--output', '-o', type=str, help='Chemin du fichier Excel à générer')
    parser.add_argument('--logs-dir', type=str, help='Répertoire des fichiers de logs (par défaut: ./logs)')
    parser.add_argument('--outputs-dir', type=str, help='Répertoire des fichiers de rapports (par défaut: ./outputs)')
    parser.add_argument('--verbose', '-v', action='store_true', help='Afficher les messages de débogage')
    
    return parser.parse_args()


def main():
    """
    Fonction principale du script.
    """
    # Analyser les arguments de la ligne de commande
    args = parse_arguments()
    
    # Configurer le niveau de logging
    if args.verbose:
        logger.setLevel(logging.DEBUG)
    
    # Créer le générateur de carnet de bord
    generator = TradeJournalGenerator(output_file=args.output)
    
    # Configurer les répertoires personnalisés si spécifiés
    if args.logs_dir:
        generator.logs_dir = args.logs_dir
    if args.outputs_dir:
        generator.outputs_dir = args.outputs_dir
    
    # Collecter les données
    logger.info("Démarrage de la collecte des données...")
    generator.collect_data()
    
    # Générer le rapport Excel
    logger.info("Génération du rapport Excel...")
    success = generator.generate_excel_report()
    
    if success:
        logger.info(f"Rapport généré avec succès : {generator.output_file}")
    else:
        logger.error("Erreur lors de la génération du rapport")
        return 1
    
    return 0


if __name__ == "__main__":
    sys.exit(main())

